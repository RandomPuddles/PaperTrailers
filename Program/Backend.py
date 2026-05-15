"""
Backend.py

Backend for a Streamlit AI Meeting Task + Research Assistant using Gemini.

What it does:
1. Takes a meeting transcript from your Streamlit UI.
2. Uses Gemini to extract:
   - participants
   - tasks
   - assignments
   - research queries
3. Downloads/searches the Kaggle arXiv dataset.
4. Saves the result to an Excel file.
5. Returns Python lists/dictionaries that Streamlit can display as tables.

Install required packages:

    pip install streamlit google-genai pydantic openpyxl python-dotenv kagglehub pandas

Create a .env file in the same folder:

    GEMINI_API_KEY=your_gemini_api_key_here
    GEMINI_MODEL=gemini-2.5-flash

Optional:

    ARXIV_JSONL_PATH=C:/path/to/arxiv-metadata-oai-snapshot.json

If ARXIV_JSONL_PATH is not provided, the program will try to download
the Kaggle arXiv dataset from the internet using kagglehub.

If Kaggle asks for credentials, create a Kaggle API token from your Kaggle
account settings and place kaggle.json in:

    C:/Users/YOUR_USERNAME/.kaggle/kaggle.json
"""

from __future__ import annotations

import heapq
import json
import os
import re
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import kagglehub
from dotenv import load_dotenv
from google import genai
from google.genai import types
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from pydantic import BaseModel, Field


# ============================================================
# Gemini structured output models
# ============================================================

class TaskItem(BaseModel):
    task_id: int = Field(description="Unique task number, starting at 1.")
    description: str = Field(description="Clean task/action item extracted from the meeting.")
    suggested_owner: str = Field(
        default="",
        description="The person who seems responsible for this task. Empty string if unclear."
    )
    evidence: str = Field(
        description="Short quote or explanation from the transcript supporting this task."
    )


class AssignmentItem(BaseModel):
    person: str = Field(description="Team member assigned to one or more tasks.")
    task_ids: List[int] = Field(description="Task IDs assigned to this person.")
    reason: str = Field(description="Why this person was assigned these tasks.")
    research_query: str = Field(
        description="Search query for finding helpful arXiv research papers."
    )


class MeetingAnalysis(BaseModel):
    participants: List[str] = Field(description="People detected in the transcript.")
    tasks: List[TaskItem] = Field(description="Tasks extracted from the transcript.")
    assignments: List[AssignmentItem] = Field(description="Task assignments by person.")


# ============================================================
# arXiv paper result model
# ============================================================

class PaperResult(BaseModel):
    person: str
    query: str
    rank: int
    score: float
    arxiv_id: str
    title: str
    authors: str
    categories: str
    abstract: str
    url: str


# ============================================================
# Kaggle dataset helper
# ============================================================

def get_arxiv_dataset_path() -> Path:
    """
    Downloads the Kaggle arXiv dataset if needed and returns the path
    to arxiv-metadata-oai-snapshot.json.

    Dataset handle:
    Cornell-University/arxiv
    """

    dataset_dir = Path(
        kagglehub.dataset_download("Cornell-University/arxiv")
    )

    possible_files = list(dataset_dir.rglob("arxiv-metadata-oai-snapshot.json"))

    if not possible_files:
        possible_files = list(dataset_dir.rglob("*.json"))

    if not possible_files:
        raise FileNotFoundError(
            "The Kaggle dataset downloaded, but no JSON metadata file was found."
        )

    return possible_files[0]


# ============================================================
# Main backend class
# ============================================================

class ResearchAssistantBackend:
    def __init__(
        self,
        arxiv_jsonl_path: Optional[str] = None,
        api_key: Optional[str] = None,
        model: Optional[str] = None,
    ):
        """
        Streamlit-friendly backend constructor.

        You can pass api_key directly from the Streamlit sidebar,
        or use a .env file:

            GEMINI_API_KEY=your_key_here
            GEMINI_MODEL=gemini-2.5-flash
        """

        load_dotenv()

        api_key = api_key or os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")

        if not api_key:
            raise ValueError(
                "GEMINI_API_KEY is missing.\n\n"
                "Create a .env file with:\n"
                "GEMINI_API_KEY=your_gemini_api_key_here\n\n"
                "or paste your Gemini API key into the Streamlit sidebar."
            )

        self.client = genai.Client(api_key=api_key)
        self.model = model or os.getenv("GEMINI_MODEL", "gemini-2.5-flash")

        if arxiv_jsonl_path:
            self.arxiv_jsonl_path = Path(arxiv_jsonl_path)
        elif os.getenv("ARXIV_JSONL_PATH"):
            self.arxiv_jsonl_path = Path(os.getenv("ARXIV_JSONL_PATH", ""))
        else:
            self.arxiv_jsonl_path = get_arxiv_dataset_path()

        if not self.arxiv_jsonl_path.exists():
            raise FileNotFoundError(
                f"Could not find arXiv dataset file:\n{self.arxiv_jsonl_path}"
            )

    # ========================================================
    # Step 1: Analyze transcript with Gemini
    # ========================================================

    def analyze_transcript(self, transcript: str) -> MeetingAnalysis:
        """
        Sends the transcript to Gemini and receives structured JSON output.
        """

        if not transcript.strip():
            raise ValueError("Transcript is empty.")

        system_prompt = """
You are the backend intelligence for an AI productivity and research assistant.

Your job:
1. Identify all meeting participants.
2. Extract clean tasks/action items from the transcript.
3. Assign tasks to people based on what they volunteered to do.
4. Generate a research query for each person based on their assigned work.

Rules:
- Do not invent people.
- Do not invent unrelated tasks.
- If someone says "I can", "I'll", "I will", or similar, treat that as strong evidence.
- Combine duplicate or highly related tasks.
- Keep task descriptions clear and direct.
- Research queries should be useful for searching arXiv academic paper metadata.
- Research queries should be specific, not vague.
- If no tasks are found, return empty lists for tasks and assignments.
"""

        response = self.client.models.generate_content(
            model=self.model,
            contents=transcript,
            config=types.GenerateContentConfig(
                system_instruction=system_prompt,
                temperature=0.2,
                max_output_tokens=4096,
                response_mime_type="application/json",
                response_json_schema=MeetingAnalysis.model_json_schema(),
            ),
        )

        json_text = self._extract_json_string(response.text)

        try:
            return MeetingAnalysis.model_validate_json(json_text)
        except Exception as error:
            raise ValueError(
                "Gemini returned output, but it could not be parsed into the expected structure.\n\n"
                f"Raw Gemini output:\n{response.text}\n\n"
                f"Original parsing error:\n{error}"
            ) from error

    # ========================================================
    # Step 2: Search arXiv dataset
    # ========================================================

    def search_arxiv_for_assignments(
        self,
        analysis: MeetingAnalysis,
        top_n_per_person: int = 3,
        max_records: Optional[int] = 100_000,
    ) -> List[PaperResult]:
        """
        Searches the arXiv metadata dataset for papers matching each person's
        research query.

        top_n_per_person:
            Number of paper recommendations for each person.

        max_records:
            100_000 is faster for testing.
            None searches the whole dataset, but it may be slower.
        """

        if top_n_per_person <= 0:
            raise ValueError("top_n_per_person must be greater than 0.")

        if not analysis.assignments:
            return []

        query_by_person: Dict[str, str] = {
            assignment.person: assignment.research_query
            for assignment in analysis.assignments
            if assignment.research_query.strip()
        }

        if not query_by_person:
            return []

        query_terms_by_person: Dict[str, List[str]] = {
            person: self._tokenize(query)
            for person, query in query_by_person.items()
        }

        best_matches: Dict[str, List[Tuple[float, int, dict]]] = {
            person: []
            for person in query_by_person
        }

        with self.arxiv_jsonl_path.open("r", encoding="utf-8") as file:
            for index, line in enumerate(file):
                if max_records is not None and index >= max_records:
                    break

                try:
                    record = json.loads(line)
                except json.JSONDecodeError:
                    continue

                for person, terms in query_terms_by_person.items():
                    if not terms:
                        continue

                    score = self._score_record(record, terms)

                    if score <= 0:
                        continue

                    heap = best_matches[person]
                    heap_item = (score, index, record)

                    if len(heap) < top_n_per_person:
                        heapq.heappush(heap, heap_item)
                    else:
                        heapq.heappushpop(heap, heap_item)

        results: List[PaperResult] = []

        for person, heap in best_matches.items():
            sorted_matches = sorted(heap, key=lambda item: item[0], reverse=True)

            for rank, item in enumerate(sorted_matches, start=1):
                score, _, record = item

                arxiv_id = self._clean_text(record.get("id", ""))
                title = self._clean_text(record.get("title", ""))
                authors = self._clean_text(record.get("authors", ""))
                categories = self._clean_text(record.get("categories", ""))
                abstract = self._clean_text(record.get("abstract", ""))

                results.append(
                    PaperResult(
                        person=person,
                        query=query_by_person[person],
                        rank=rank,
                        score=round(score, 2),
                        arxiv_id=arxiv_id,
                        title=title,
                        authors=authors,
                        categories=categories,
                        abstract=abstract,
                        url=f"https://arxiv.org/abs/{arxiv_id}",
                    )
                )

        return results

    # ========================================================
    # Step 3: Export to Excel
    # ========================================================

    def export_to_excel(
        self,
        analysis: MeetingAnalysis,
        papers: List[PaperResult],
        output_path: str,
    ) -> str:
        """
        Saves tasks, assignments, participants, and paper recommendations
        into an Excel workbook.
        """

        wb = Workbook()

        # ----------------------------
        # Sheet 1: Tasks
        # ----------------------------

        ws_tasks = wb.active
        ws_tasks.title = "Tasks"

        ws_tasks.append([
            "Task ID",
            "Description",
            "Suggested Owner",
            "Evidence",
        ])

        for task in analysis.tasks:
            ws_tasks.append([
                task.task_id,
                self._excel_safe(task.description),
                self._excel_safe(task.suggested_owner),
                self._excel_safe(task.evidence),
            ])

        # ----------------------------
        # Sheet 2: Assignments
        # ----------------------------

        ws_assignments = wb.create_sheet("Assignments")

        ws_assignments.append([
            "Person",
            "Task IDs",
            "Reason",
            "Research Query",
        ])

        for assignment in analysis.assignments:
            ws_assignments.append([
                self._excel_safe(assignment.person),
                ", ".join(str(task_id) for task_id in assignment.task_ids),
                self._excel_safe(assignment.reason),
                self._excel_safe(assignment.research_query),
            ])

        # ----------------------------
        # Sheet 3: Research Recommendations
        # ----------------------------

        ws_papers = wb.create_sheet("Research Recommendations")

        ws_papers.append([
            "Person",
            "Search Query",
            "Rank",
            "Score",
            "arXiv ID",
            "Title",
            "Authors",
            "Categories",
            "Abstract",
            "URL",
        ])

        for paper in papers:
            ws_papers.append([
                self._excel_safe(paper.person),
                self._excel_safe(paper.query),
                paper.rank,
                paper.score,
                self._excel_safe(paper.arxiv_id),
                self._excel_safe(paper.title),
                self._excel_safe(paper.authors),
                self._excel_safe(paper.categories),
                self._excel_safe(paper.abstract),
                self._excel_safe(paper.url),
            ])

        # ----------------------------
        # Sheet 4: Participants
        # ----------------------------

        ws_people = wb.create_sheet("Participants")

        ws_people.append(["Participants"])

        for participant in analysis.participants:
            ws_people.append([
                self._excel_safe(participant)
            ])

        self._format_workbook(wb)

        output_path = str(Path(output_path).resolve())
        wb.save(output_path)

        return output_path

    # ========================================================
    # Main function for Streamlit frontend to call
    # ========================================================

    def process_transcript(
        self,
        transcript: str,
        output_excel_path: str = "meeting_research_output.xlsx",
        top_n_per_person: int = 3,
        max_arxiv_records: Optional[int] = 100_000,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> Dict[str, Any]:
        """
        This is the main method your Streamlit frontend should call.

        Returns:
            Dictionary containing:
            - excel_path
            - participants
            - tasks
            - assignments
            - research_recommendations
        """

        self._send_progress(progress_callback, "Analyzing transcript with Gemini...")
        analysis = self.analyze_transcript(transcript)

        self._send_progress(progress_callback, "Searching arXiv research papers...")
        papers = self.search_arxiv_for_assignments(
            analysis=analysis,
            top_n_per_person=top_n_per_person,
            max_records=max_arxiv_records,
        )

        self._send_progress(progress_callback, "Saving Excel file...")
        saved_path = self.export_to_excel(
            analysis=analysis,
            papers=papers,
            output_path=output_excel_path,
        )

        self._send_progress(progress_callback, "Done.")

        return {
            "excel_path": saved_path,
            "participants": self.participants_to_rows(analysis),
            "tasks": self.tasks_to_rows(analysis),
            "assignments": self.assignments_to_rows(analysis),
            "research_recommendations": self.papers_to_rows(papers),
            "raw_analysis": analysis,
            "raw_papers": papers,
        }

    # ========================================================
    # Streamlit-friendly row converters
    # ========================================================

    def participants_to_rows(self, analysis: MeetingAnalysis) -> List[Dict[str, str]]:
        return [
            {
                "Participant": participant,
            }
            for participant in analysis.participants
        ]

    def tasks_to_rows(self, analysis: MeetingAnalysis) -> List[Dict[str, Any]]:
        return [
            {
                "Task ID": task.task_id,
                "Description": task.description,
                "Suggested Owner": task.suggested_owner,
                "Evidence": task.evidence,
            }
            for task in analysis.tasks
        ]

    def assignments_to_rows(self, analysis: MeetingAnalysis) -> List[Dict[str, Any]]:
        return [
            {
                "Person": assignment.person,
                "Task IDs": ", ".join(str(task_id) for task_id in assignment.task_ids),
                "Reason": assignment.reason,
                "Research Query": assignment.research_query,
            }
            for assignment in analysis.assignments
        ]

    def papers_to_rows(self, papers: List[PaperResult]) -> List[Dict[str, Any]]:
        return [
            {
                "Person": paper.person,
                "Search Query": paper.query,
                "Rank": paper.rank,
                "Score": paper.score,
                "arXiv ID": paper.arxiv_id,
                "Title": paper.title,
                "Authors": paper.authors,
                "Categories": paper.categories,
                "Abstract": paper.abstract,
                "URL": paper.url,
            }
            for paper in papers
        ]

    # ========================================================
    # Helper functions
    # ========================================================

    def _send_progress(
        self,
        progress_callback: Optional[Callable[[str], None]],
        message: str,
    ) -> None:
        if progress_callback:
            progress_callback(message)

    def _tokenize(self, text: str) -> List[str]:
        """
        Breaks a research query into important searchable terms.
        """

        stop_words = {
            "the", "and", "for", "with", "that", "this", "from", "into",
            "how", "what", "when", "where", "why", "are", "can", "use",
            "using", "over", "time", "project", "research", "paper",
            "papers", "article", "articles", "study", "studies", "based",
            "related", "approach", "method", "methods", "system", "systems",
            "model", "models", "analysis", "application", "applications",
            "implementation", "implementing", "design", "designing"
        }

        words = re.findall(r"[a-zA-Z][a-zA-Z0-9\-]+", text.lower())

        return [
            word
            for word in words
            if len(word) >= 3 and word not in stop_words
        ]

    def _score_record(self, record: dict, query_terms: List[str]) -> float:
        """
        Scores one arXiv paper based on how well it matches a research query.
        Higher score = better match.
        """

        title = str(record.get("title", "")).lower()
        abstract = str(record.get("abstract", "")).lower()
        categories = str(record.get("categories", "")).lower()

        score = 0.0

        for term in query_terms:
            score += title.count(term) * 10
            score += categories.count(term) * 5
            score += abstract.count(term) * 2

        return score

    def _clean_text(self, value: object) -> str:
        """
        Cleans text from the dataset so Excel cells look nicer.
        """

        text = str(value or "")
        text = re.sub(r"\s+", " ", text).strip()

        # Excel max cell length is 32,767 characters.
        if len(text) > 32_000:
            text = text[:32_000] + "..."

        return text

    def _excel_safe(self, value: object) -> str:
        """
        Prevents Excel formula injection.

        If a cell starts with =, +, -, or @, Excel may treat it as a formula.
        Prefixing with an apostrophe makes Excel treat it as plain text.
        """

        text = self._clean_text(value)

        if text.startswith(("=", "+", "-", "@")):
            text = "'" + text

        return text

    def _format_workbook(self, wb: Workbook) -> None:
        """
        Makes the Excel workbook easier to read.
        """

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )

            for column_cells in ws.columns:
                column_letter = column_cells[0].column_letter
                header = str(column_cells[0].value or "").lower()

                if "abstract" in header:
                    width = 80
                elif "description" in header or "evidence" in header or "reason" in header:
                    width = 50
                elif "url" in header:
                    width = 45
                elif "title" in header:
                    width = 55
                else:
                    width = 25

                ws.column_dimensions[column_letter].width = width

            ws.freeze_panes = "A2"

    def _extract_json_string(self, text: str) -> str:
        """
        Gemini should return pure JSON because response_mime_type is application/json.
        This function also handles accidental markdown code fences just in case.
        """

        cleaned = text.strip()

        if cleaned.startswith("```"):
            cleaned = re.sub(r"^```(?:json)?", "", cleaned, flags=re.IGNORECASE).strip()
            cleaned = re.sub(r"```$", "", cleaned).strip()

        first_brace = cleaned.find("{")
        last_brace = cleaned.rfind("}")

        if first_brace != -1 and last_brace != -1 and last_brace > first_brace:
            cleaned = cleaned[first_brace:last_brace + 1]

        return cleaned


# ============================================================
# Optional sample transcript for testing backend directly
# ============================================================

SAMPLE_TRANSCRIPT = """
Ian: For our physics project, I think we should build a simulation that shows how projectile motion changes when air resistance is included.
Maya: I can research the equations for drag force and how we update velocity and position over time.
Jordan: I’ll handle the Python simulation logic and make sure the projectile path updates correctly with different masses and launch angles.
Alex: I can compare the air resistance version to the ideal no-drag version and explain the difference in the final demo.
Ian: Great, I’ll build the interface so users can enter the launch angle, starting velocity, mass, and drag coefficient.
"""


# ============================================================
# Allows quick testing with: python Backend.py
# ============================================================

if __name__ == "__main__":
    backend = ResearchAssistantBackend()

    result = backend.process_transcript(
        transcript=SAMPLE_TRANSCRIPT,
        output_excel_path="meeting_research_output.xlsx",
        top_n_per_person=3,
        max_arxiv_records=100_000,
        progress_callback=print,
    )

    print()
    print("Excel saved to:")
    print(result["excel_path"])

    print()
    print("Tasks:")
    for row in result["tasks"]:
        print(row)

    print()
    print("Assignments:")
    for row in result["assignments"]:
        print(row)

    print()
    print("Research Recommendations:")
    for row in result["research_recommendations"]:
        print(row)